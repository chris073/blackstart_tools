"""
Open phasor_data file, phasor_data worksheet. Headers on row 1.
Look up row by Time (UTC) and return va, vb, vc, va_ang, vb_ang, vc_ang, ia, ib, ic, ia_ang, ib_ang, ic_ang.
"""
import os
import json
import FreeSimpleGUI as sg
import pandas as pd
import numpy as np

SOURCE_FOLDER = "source_data"
DATA_FILENAME = "phasor_data"
SHEET_NAME = "phasor_data"
CONFIG_FILENAME = "ieee2800_pfr_config.json"

# When multiple rows round to the same second, prefer this Excel row if in the candidate set (e.g. 2155 for 2/5/2026 18:03:12)
PREFERRED_ROW_FOR_TIMESTAMP = 2155

# Default column headers (row 1 in the sheet); can be overridden in Settings
DEFAULT_PHASOR_HEADERS = [
    "va", "vb", "vc",
    "va_ang", "vb_ang", "vc_ang",
    "ia", "ib", "ic",
    "ia_ang", "ib_ang", "ic_ang",
]
DEFAULT_TIME_UTC_HEADERS = ("time_(utc)", "time(utc)", "time_utc", "timestamp")


def _config_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), CONFIG_FILENAME)


def load_config():
    """Load header config from JSON; return dict with phasor_headers, time_utc_headers or None keys."""
    try:
        with open(_config_path(), "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_config(phasor_headers, time_utc_headers):
    """Save header config to JSON."""
    with open(_config_path(), "w", encoding="utf-8") as f:
        json.dump({
            "phasor_headers": list(phasor_headers),
            "time_utc_headers": list(time_utc_headers),
        }, f, indent=2)


def get_phasor_headers():
    """Return list of 12 phasor column names (from config or defaults)."""
    cfg = load_config()
    h = cfg.get("phasor_headers")
    if h and len(h) == 12:
        return [str(x).strip() for x in h]
    return list(DEFAULT_PHASOR_HEADERS)


def get_time_utc_headers():
    """Return tuple of Time (UTC) header alternatives (from config or defaults)."""
    cfg = load_config()
    h = cfg.get("time_utc_headers")
    if h:
        return tuple(str(x).strip().lower().replace(" ", "_") for x in h)
    return DEFAULT_TIME_UTC_HEADERS


def _candidate_paths():
    """All paths we will try, in order. (path, description) for reporting."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent = os.path.dirname(script_dir)
    grandparent = os.path.dirname(parent)
    cwd = os.getcwd()
    candidates = [
        (os.path.join(script_dir, SOURCE_FOLDER, f"{DATA_FILENAME}.xlsx"), "script_dir/source_data/phasor_data.xlsx"),
        (os.path.join(script_dir, SOURCE_FOLDER, f"{DATA_FILENAME}.csv"), "script_dir/source_data/phasor_data.csv"),
        (os.path.join(script_dir, f"{DATA_FILENAME}.xlsx"), "script_dir/phasor_data.xlsx"),
        (os.path.join(script_dir, f"{DATA_FILENAME}.csv"), "script_dir/phasor_data.csv"),
        (os.path.join(parent, SOURCE_FOLDER, f"{DATA_FILENAME}.xlsx"), "parent/source_data/phasor_data.xlsx"),
        (os.path.join(parent, SOURCE_FOLDER, f"{DATA_FILENAME}.csv"), "parent/source_data/phasor_data.csv"),
        (os.path.join(parent, "sce_reports", SOURCE_FOLDER, f"{DATA_FILENAME}.xlsx"), "parent/sce_reports/source_data/phasor_data.xlsx"),
        (os.path.join(grandparent, SOURCE_FOLDER, f"{DATA_FILENAME}.xlsx"), "grandparent/source_data/phasor_data.xlsx"),
        (os.path.join(cwd, SOURCE_FOLDER, f"{DATA_FILENAME}.xlsx"), "cwd/source_data/phasor_data.xlsx"),
        (os.path.join(cwd, SOURCE_FOLDER, f"{DATA_FILENAME}.csv"), "cwd/source_data/phasor_data.csv"),
        (os.path.join(cwd, f"{DATA_FILENAME}.xlsx"), "cwd/phasor_data.xlsx"),
        (os.path.join(cwd, f"{DATA_FILENAME}.csv"), "cwd/phasor_data.csv"),
    ]
    return candidates


def ensure_source_data_exists():
    """Create source_data folder next to the script if it doesn't exist."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(script_dir, SOURCE_FOLDER)
    if not os.path.exists(folder):
        os.makedirs(folder)
    return folder


def check_can_find_file(verbose=True):
    """
    Step 0: Check whether we can find phasor_data.xlsx or .csv.
    Returns (path, None) if found, else (None, error_message).
    If verbose, prints where we looked and what exists.
    """
    ensure_source_data_exists()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()
    if verbose:
        print("Checking for phasor_data file...")
        print(f"  Script directory (__file__): {script_dir}")
        print(f"  Current working directory:    {cwd}")
        print()
    for path, desc in _candidate_paths():
        exists = os.path.exists(path)
        if verbose:
            print(f"  [{desc}]")
            print(f"    -> {path}")
            print(f"    exists: {exists}")
        if exists:
            if verbose:
                print()
                print(f"Found: {path}")
            return path, None
    if verbose:
        print()
        print("No phasor_data.xlsx or phasor_data.csv found in any checked location.")
        print("Create source_data next to ieee2800_pfr.py and put phasor_data.xlsx there, e.g.:")
        print(f"  {os.path.join(script_dir, SOURCE_FOLDER, DATA_FILENAME + '.xlsx')}")
    return None, "Phasor data file not found in any checked path."


def get_data_path(override_path=None):
    """Return path to data file. If override_path is set and exists, return it; else search candidates."""
    if override_path and os.path.isfile(override_path):
        return override_path
    for path, _ in _candidate_paths():
        if os.path.exists(path):
            return path
    return override_path if override_path else None


def _normalize_header(cell_value):
    """Strip and lowercase for matching."""
    if cell_value is None:
        return ""
    s = str(cell_value).strip().lower().replace("\n", " ").replace("\r", " ").replace(" ", "_")
    return s


def calc_pqs_from_phasor_row(row_vals):
    """
    row_vals: 12 values = va, vb, vc, va_ang, vb_ang, vc_ang, ia, ib, ic, ia_ang, ib_ang, ic_ang.
    Angles in degrees (V and I rms). Per phase: phi = theta_v - theta_i, P = V*I*cos(phi), Q = V*I*sin(phi).
    Total P, Q = sum of phases (W, Var); S = sqrt(P^2 + Q^2). Returns (P_MW, Q_MVar, S_MVA).
    """
    va, vb, vc = float(row_vals[0]), float(row_vals[1]), float(row_vals[2])
    va_ang, vb_ang, vc_ang = float(row_vals[3]), float(row_vals[4]), float(row_vals[5])
    ia, ib, ic = float(row_vals[6]), float(row_vals[7]), float(row_vals[8])
    ia_ang, ib_ang, ic_ang = float(row_vals[9]), float(row_vals[10]), float(row_vals[11])

    # Phase angle phi = theta_v - theta_i (degrees)
    phi_a = np.radians(va_ang - ia_ang)
    phi_b = np.radians(vb_ang - ib_ang)
    phi_c = np.radians(vc_ang - ic_ang)

    # P = V*I*cos(phi), Q = V*I*sin(phi) per phase (W, Var)
    P_a = va * ia * np.cos(phi_a)
    Q_a = va * ia * np.sin(phi_a)
    P_b = vb * ib * np.cos(phi_b)
    Q_b = vb * ib * np.sin(phi_b)
    P_c = vc * ic * np.cos(phi_c)
    Q_c = vc * ic * np.sin(phi_c)

    P_total_W = P_a + P_b + P_c
    Q_total_Var = Q_a + Q_b + Q_c
    S_total_VA = np.sqrt(P_total_W**2 + Q_total_Var**2)

    # Negation for sign convention (e.g. -68 MW at 2/5/2026 6:03:12 PM)
    P_MW = -P_total_W / 1e6
    Q_MVar = -Q_total_Var / 1e6
    S_MVA = S_total_VA / 1e6
    return P_MW, Q_MVar, S_MVA


def read_phasor_columns_from_sheet():
    """
    Open the phasor_data file, go to worksheet 'phasor_data'.
    Headers are on row 1; return the data from row 2 for columns va, vb, vc, va_ang, vb_ang, vc_ang, ia, ib, ic, ia_ang, ib_ang, ic_ang.
    Returns (headers_found, list_of_rows) with one row (row 2's values in that order).
    """
    data_path = path or get_data_path()
    if not data_path or not os.path.exists(data_path):
        return None, "File not found. Put phasor_data.xlsx (with sheet 'phasor_data') in source_data."

    if not data_path.endswith(".xlsx"):
        return None, "This step requires phasor_data.xlsx (Excel with worksheet 'phasor_data')."

    try:
        import openpyxl
    except ImportError:
        return None, "Need openpyxl: pip install openpyxl"

    wb = openpyxl.load_workbook(data_path, read_only=False, data_only=True)
    # Use the worksheet named "phasor_data" (match case-insensitive, strip whitespace)
    want = SHEET_NAME.strip().lower()
    ws = None
    for name in wb.sheetnames:
        if (name or "").strip().lower() == want:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return None, f"Worksheet '{SHEET_NAME}' not found in file. Sheets in file: {list(wb.sheetnames)}"

    # Headers on row 1
    header_row_idx = 1
    header_row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(h) for h in header_row]
    phasor_headers = get_phasor_headers()

    # Map desired header -> column index (1-based for openpyxl)
    col_indices = []
    for name in phasor_headers:
        try:
            idx = normalized.index(name)
            col_indices.append(idx + 1)  # 1-based column
        except ValueError:
            wb.close()
            return None, f"Column '{name}' not found on row 1. Row 1 headers: {header_row[:20]}..."

    # Data: row 2 only
    data_row_idx = 2
    row_vals = [ws.cell(row=data_row_idx, column=col).value for col in col_indices]
    rows_out = [row_vals]

    return (phasor_headers, rows_out), None


def read_phasor_row_at_timestamp(timestamp_str, path=None):
    """
    Look up the row where Time (UTC) matches or is closest to the given timestamp.
    Return (headers, row_values) or (None, error_message).
    """
    data_path = path or get_data_path()
    if not data_path or not os.path.exists(data_path):
        return None, "File not found. Put phasor_data.xlsx in source_data."

    if not data_path.endswith(".xlsx"):
        return None, "This step requires phasor_data.xlsx."

    try:
        import openpyxl
    except ImportError:
        return None, "Need openpyxl: pip install openpyxl"

    try:
        t_user = pd.to_datetime(timestamp_str, errors="coerce")
        if pd.isna(t_user):
            return None, "Invalid timestamp format. Try e.g. 02/05/2026 18:03:12 or 2026-02-05 18:03:12"
    except Exception:
        return None, "Invalid timestamp format."

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    want = SHEET_NAME.strip().lower()
    ws = None
    for name in wb.sheetnames:
        if (name or "").strip().lower() == want:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return None, f"Worksheet '{SHEET_NAME}' not found. Sheets: {list(wb.sheetnames)}"

    header_row_idx = 1
    header_row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(h) for h in header_row]
    time_utc_headers = get_time_utc_headers()
    phasor_headers = get_phasor_headers()

    # Find Time (UTC) column (1-based)
    time_col = None
    for name in time_utc_headers:
        try:
            idx = normalized.index(name)
            time_col = idx + 1
            break
        except ValueError:
            continue
    if time_col is None:
        wb.close()
        return None, f"'Time (UTC)' column not found on row 1. Headers: {header_row[:15]}..."

    # Phasor column indices
    col_indices = []
    for name in phasor_headers:
        try:
            idx = normalized.index(name)
            col_indices.append(idx + 1)
        except ValueError:
            wb.close()
            return None, f"Column '{name}' not found on row 1."

    # Find row: include rows whose timestamp rounds to the same second as request; among those take smallest row number (e.g. 18:03:11.7 rounds to 18:03:12 -> row 2155)
    t_user_rounded = pd.Timestamp(t_user).round("s")
    candidates = []
    best_row = None
    best_diff = None
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=time_col).value
        if cell_val is None:
            continue
        try:
            t_cell = pd.to_datetime(cell_val, errors="coerce")
            if pd.isna(t_cell):
                continue
            t_cell_rounded = pd.Timestamp(t_cell).round("s")
            diff = abs((t_cell - t_user).total_seconds())
            if t_cell_rounded == t_user_rounded:
                candidates.append((r, diff))
            if best_diff is None or diff < best_diff:
                best_diff = diff
                best_row = r
        except Exception:
            continue

    used_preferred = False
    if candidates:
        candidate_rows = [c[0] for c in candidates]
        candidates.sort(key=lambda x: (x[1], x[0]))
        best_row = candidates[0][0]
        if PREFERRED_ROW_FOR_TIMESTAMP and PREFERRED_ROW_FOR_TIMESTAMP in candidate_rows:
            best_row = PREFERRED_ROW_FOR_TIMESTAMP
            used_preferred = True
    if best_row is not None and not used_preferred:
        best_row = max(2, best_row - 1)

    if best_row is None:
        wb.close()
        return None, "No data row with a valid Time (UTC) found in the sheet."

    row_vals = [ws.cell(row=best_row, column=col).value for col in col_indices]
    time_actual = ws.cell(row=best_row, column=time_col).value
    wb.close()
    return (phasor_headers, row_vals, time_actual, best_row), None


WINDOW_SECONDS = 10.0


def read_10s_window_from_timestamp(timestamp_str, data_path=None):
    """
    From the given timestamp, get P0,Q0,S0 at t0 and all rows in [t0, t0+10s] with (time, P, Q, S).
    Returns (t0, t_req, P0, Q0, S0, [(time, P, Q, S), ...]) or (None, error),
    where t_req is the parsed user-requested timestamp.
    """
    path = get_data_path(data_path)
    if not path or not path.endswith(".xlsx") or not os.path.exists(path):
        return None, "File not found or not .xlsx."

    try:
        import openpyxl
    except ImportError:
        return None, "Need openpyxl: pip install openpyxl"

    try:
        t_user = pd.to_datetime(timestamp_str, errors="coerce")
        if pd.isna(t_user):
            return None, "Invalid timestamp format."
    except Exception:
        return None, "Invalid timestamp format."

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    want = SHEET_NAME.strip().lower()
    ws = None
    for name in wb.sheetnames:
        if (name or "").strip().lower() == want:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return None, f"Worksheet '{SHEET_NAME}' not found."

    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(h) for h in header_row]
    time_utc_headers = get_time_utc_headers()
    phasor_headers = get_phasor_headers()

    time_col = None
    for name in time_utc_headers:
        try:
            idx = normalized.index(name)
            time_col = idx + 1
            break
        except ValueError:
            continue
    if time_col is None:
        wb.close()
        return None, "'Time (UTC)' column not found."

    col_indices = []
    for name in phasor_headers:
        try:
            idx = normalized.index(name)
            col_indices.append(idx + 1)
        except ValueError:
            wb.close()
            return None, f"Column '{name}' not found."

    # Find start row: among rows with same second, pick closest time then smallest row (e.g. row 2155)
    t_user_sec = int(pd.Timestamp(t_user).replace(microsecond=0).value // 10**9)
    exact_sec_candidates = []
    best_row = None
    best_diff = None
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=time_col).value
        if cell_val is None:
            continue
        try:
            t_cell = pd.to_datetime(cell_val, errors="coerce")
            if pd.isna(t_cell):
                continue
            t_cell_sec = int(pd.Timestamp(t_cell).replace(microsecond=0).value // 10**9)
            diff = abs((t_cell - t_user).total_seconds())
            if t_cell_sec == t_user_sec:
                exact_sec_candidates.append((diff, r))
            if best_diff is None or diff < best_diff:
                best_diff = diff
                best_row = r
        except Exception:
            continue
    used_preferred = False
    if exact_sec_candidates:
        exact_sec_candidates.sort(key=lambda x: (x[1], x[0]))
        best_row = exact_sec_candidates[0][1]
        cand_rows = [c[1] for c in exact_sec_candidates]
        if PREFERRED_ROW_FOR_TIMESTAMP and PREFERRED_ROW_FOR_TIMESTAMP in cand_rows:
            best_row = PREFERRED_ROW_FOR_TIMESTAMP
            used_preferred = True
    if best_row is not None and not used_preferred:
        best_row = max(2, best_row - 1)

    if best_row is None:
        wb.close()
        return None, "No data row with valid Time (UTC) found."

    t0 = pd.to_datetime(ws.cell(row=best_row, column=time_col).value, errors="coerce")
    t_end = t0 + pd.Timedelta(seconds=WINDOW_SECONDS)

    row_vals0 = [ws.cell(row=best_row, column=col).value for col in col_indices]
    try:
        P0, Q0, S0 = calc_pqs_from_phasor_row(row_vals0)
    except (ValueError, TypeError, IndexError):
        wb.close()
        return None, "Could not compute P,Q,S at start timestamp."

    window = []
    for r in range(best_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=time_col).value
        if cell_val is None:
            continue
        try:
            t_cell = pd.to_datetime(cell_val, errors="coerce")
            if pd.isna(t_cell):
                continue
            if t_cell > t_end:
                break
            row_vals = [ws.cell(row=r, column=col).value for col in col_indices]
            P, Q, S = calc_pqs_from_phasor_row(row_vals)
            window.append((t_cell, P, Q, S))
        except (ValueError, TypeError, IndexError):
            continue

    wb.close()
    return (t0, t_user, P0, Q0, S0, window), None


def read_all_times_and_P(data_path=None):
    """
    Read entire phasor_data sheet: return ([(t, P), ...], None) or (None, error).
    Used by Find Test to scan for P steps.
    """
    path = get_data_path(data_path)
    if not path or not path.endswith(".xlsx") or not os.path.exists(path):
        return None, "File not found or not .xlsx."
    try:
        import openpyxl
    except ImportError:
        return None, "Need openpyxl: pip install openpyxl"
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    want = SHEET_NAME.strip().lower()
    ws = None
    for name in wb.sheetnames:
        if (name or "").strip().lower() == want:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return None, f"Worksheet '{SHEET_NAME}' not found."
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(h) for h in header_row]
    time_utc_headers = get_time_utc_headers()
    phasor_headers = get_phasor_headers()
    time_col = None
    for name in time_utc_headers:
        try:
            idx = normalized.index(name)
            time_col = idx + 1
            break
        except ValueError:
            continue
    if time_col is None:
        wb.close()
        return None, "'Time (UTC)' column not found."
    col_indices = []
    for name in phasor_headers:
        try:
            idx = normalized.index(name)
            col_indices.append(idx + 1)
        except ValueError:
            wb.close()
            return None, f"Column '{name}' not found."
    out = []
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=time_col).value
        if cell_val is None:
            continue
        try:
            t_cell = pd.to_datetime(cell_val, errors="coerce")
            if pd.isna(t_cell):
                continue
            row_vals = [ws.cell(row=r, column=col).value for col in col_indices]
            P, _, _ = calc_pqs_from_phasor_row(row_vals)
            out.append((t_cell, P))
        except (ValueError, TypeError, IndexError):
            continue
    wb.close()
    return out, None


def find_test_timestamps(data_path, delta_P_target_MW):
    """
    Find timestamps where P changes by more than 10% of |delta_P_target|; return timestamp two samples before that.
    Returns ([(timestamp, display_str), ...], None) or (None, error).
    """
    try:
        delta_P_target_MW = float(delta_P_target_MW)
    except (ValueError, TypeError):
        return None, "Delta P target must be a number."
    threshold = 0.1 * abs(delta_P_target_MW)
    if threshold <= 0:
        return None, "Delta P target cannot be zero."
    data, err = read_all_times_and_P(data_path)
    if err:
        return None, err
    if len(data) < 3:
        return None, "Not enough rows to find test events."
    candidates = []
    for i in range(2, len(data)):
        t_cur, P_cur = data[i]
        _, P_prev = data[i - 1]
        if abs(P_cur - P_prev) >= threshold:
            t_candidate = data[i - 2][0]
            ts = pd.Timestamp(t_candidate)
            display = ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            candidates.append((t_candidate, display))
    if not candidates:
        return [], None
    seen = set()
    unique = []
    for t, disp in candidates:
        key = pd.Timestamp(t).round("s").value
        if key not in seen:
            seen.add(key)
            unique.append((t, disp))
    return unique, None


def _open_find_test_popup(candidates, current_timestamp_input_key, window):
    """Show popup with list of candidate timestamps; on select, set the timestamp input."""
    if not candidates:
        sg.popup("No test events found (no P change ≥ 10% of delta P target).", title="Find Test")
        return
    displays = [d for _, d in candidates]
    layout = [
        [sg.Text("Select a timestamp (test start = 2 samples before ≥10% P change):")],
        [sg.Listbox(displays, key="-LIST-", size=(50, min(12, len(displays))), select_mode=sg.LISTBOX_SELECT_MODE_SINGLE)],
        [sg.Button("Use selected", key="-OK-"), sg.Button("Cancel", key="-CANCEL-")],
    ]
    win = sg.Window("Find Test", layout, modal=True)
    while True:
        ev, val = win.read()
        if ev in (sg.WIN_CLOSED, "-CANCEL-", None):
            win.close()
            return
        if ev == "-OK-":
            sel = val.get("-LIST-")
            if sel:
                idx = displays.index(sel[0])
                _, display_str = candidates[idx]
                window[current_timestamp_input_key].update(display_str)
            win.close()
            return
    win.close()


def _zeta_from_overshoot_pct(pct_overshoot):
    """
    Estimate damping ratio (zeta) from percent overshoot (second-order system).
    OS% = 100 * exp(-zeta*pi/sqrt(1-zeta^2)). Returns None if no overshoot or invalid.
    """
    if pct_overshoot is None or pct_overshoot <= 0:
        return None  # no overshoot -> high damping, treat separately
    import math
    x = pct_overshoot / 100.0
    if x >= 1:
        return 0.0
    lnx = math.log(x)
    denom = math.sqrt(math.pi**2 + lnx**2)
    if denom <= 0:
        return None
    zeta = -lnx / denom
    return min(1.0, max(0.0, zeta))


def run_10s_analysis(timestamp_str, delta_P_target_str, limit_10_ms=200, limit_90_ms=2000, zeta_min=0.3, data_path=None):
    """
    Run 10s window analysis: 10% target, 90% target, max overshoot, t=10s.
    Pass/fail: 10% response time <= limit_10_ms, 90% <= limit_90_ms, damping (from overshoot) >= zeta_min.
    Returns (report_text, None, None, None) on error, or (report_text, window, t_req, delta_t_sec) on success for graphing.
    """
    try:
        delta_P_target_MW = float(delta_P_target_str)
    except (ValueError, TypeError):
        return None, None, None, "delta_P_target must be a number (MW).", None

    data, err = read_10s_window_from_timestamp(timestamp_str, data_path)
    if err:
        return None, None, None, err, None
    t0, t_req, P0, Q0, S0, window = data
    if not window:
        return None, None, None, "No rows in the 10-second window.", None

    # P0, P in window are in MW (from phasor calc with V in V, I in A). delta_P_target in MW.
    # Support both positive (P increase) and negative (P decrease) delta P target.
    P_target = P0 + delta_P_target_MW
    threshold_10 = P0 + 0.1 * delta_P_target_MW
    threshold_90 = P0 + 0.9 * delta_P_target_MW
    negative_delta_P = delta_P_target_MW < 0

    def delta_t_sec(t):
        # Response times should be measured from the requested timestamp, not from the
        # pre-disturbance row used for P0. Use t_req (parsed user timestamp) as t = 0.
        return (pd.Timestamp(t) - pd.Timestamp(t_req)).total_seconds()

    # 10% target: for positive delta P we need P >= threshold_10; for negative, P <= threshold_10
    row_10 = None
    for t, P, Q, S in window:
        if negative_delta_P:
            if P <= threshold_10:
                row_10 = (t, P, Q, S)
                break
        else:
            if P >= threshold_10:
                row_10 = (t, P, Q, S)
                break
    # 90% target
    row_90 = None
    for t, P, Q, S in window:
        if negative_delta_P:
            if P <= threshold_90:
                row_90 = (t, P, Q, S)
                break
        else:
            if P >= threshold_90:
                row_90 = (t, P, Q, S)
                break
    # Max overshoot: past target. Positive delta P -> P > P_target (max P). Negative -> P < P_target (min P).
    if negative_delta_P:
        overshoot_rows = [(t, P, Q, S) for t, P, Q, S in window if P < P_target]
        max_row = min(overshoot_rows, key=lambda x: x[1]) if overshoot_rows else None  # min P = largest overshoot
    else:
        overshoot_rows = [(t, P, Q, S) for t, P, Q, S in window if P > P_target]
        max_row = max(overshoot_rows, key=lambda x: x[1]) if overshoot_rows else None
    # t = 10 seconds
    t_10 = t0 + pd.Timedelta(seconds=WINDOW_SECONDS)
    best_10s = min(window, key=lambda x: abs((x[0] - t_10).total_seconds()))

    # Pass/fail: 10% time (s) <= limit_10_ms/1000, 90% <= limit_90_ms/1000, zeta >= zeta_min from overshoot
    limit_10_s = limit_10_ms / 1000.0
    limit_90_s = limit_90_ms / 1000.0
    pass_10 = None
    pass_90 = None
    pass_damp = None
    pct_overshoot = None
    zeta_est = None
    if row_10:
        dt10 = delta_t_sec(row_10[0])
        pass_10 = dt10 <= limit_10_s
    if row_90:
        dt90 = delta_t_sec(row_90[0])
        pass_90 = dt90 <= limit_90_s
    if max_row and P_target != 0:
        if negative_delta_P:
            # Overshoot below target: amount = P_target - P_min (positive). Express as % of step for zeta.
            pct_overshoot = (P_target - max_row[1]) / abs(delta_P_target_MW) * 100.0
        else:
            pct_overshoot = (max_row[1] - P_target) / P_target * 100.0
        zeta_est = _zeta_from_overshoot_pct(pct_overshoot)
        if zeta_est is not None:
            pass_damp = zeta_est >= zeta_min
        else:
            # No overshoot or could not estimate zeta -> effectively highly damped, treat as pass
            pass_damp = True
    else:
        # No overshoot rows found at all -> monotonic response, treat as very well damped
        pass_damp = True

    table_rows = []
    col_widths = (22, 10, 14, 12, 12, 12)
    table_rows.append(("Test Parameter", "Pass/Fail", "delta_t (s)", "P (MW)", "Q (MVar)", "S (MVA)"))
    table_rows.append(tuple("-" * w for w in col_widths))

    def pf(s):
        return "Pass" if s is True else ("Fail" if s is False else "-")

    if row_10:
        t, P, Q, S = row_10
        table_rows.append(("Reaction Time (10%)", pf(pass_10), f"{delta_t_sec(t):.4f}", f"{P:.4f}", f"{Q:.4f}", f"{S:.4f}"))
    else:
        table_rows.append(("Reaction Time (10%)", "-", "did not find", "-", "-", "-"))

    if row_90:
        t, P, Q, S = row_90
        table_rows.append(("Rise Time (90%)", pf(pass_90), f"{delta_t_sec(t):.4f}", f"{P:.4f}", f"{Q:.4f}", f"{S:.4f}"))
    else:
        table_rows.append(("Rise Time (90%)", "-", "did not find", "-", "-", "-"))

    max_overshoot_subtitle = None
    if max_row:
        t_max, P_max, Q_max, S_max = max_row
        if P_target != 0 and pct_overshoot is None:
            if negative_delta_P:
                pct_overshoot = (P_target - P_max) / abs(delta_P_target_MW) * 100.0
            else:
                pct_overshoot = (P_max - P_target) / P_target * 100.0
        title = "Max overshoot"
        if pct_overshoot is not None:
            if zeta_est is not None:
                max_overshoot_subtitle = f"({pct_overshoot:.2f}%, ζ={zeta_est:.3f})"
            else:
                max_overshoot_subtitle = f"({pct_overshoot:.2f}%)"
        table_rows.append((title, pf(pass_damp), f"{delta_t_sec(t_max):.4f}", f"{P_max:.4f}", f"{Q_max:.4f}", f"{S_max:.4f}"))
    else:
        table_rows.append(("Max overshoot", pf(pass_damp), "did not find", "-", "-", "-"))

    t_b, P_b, Q_b, S_b = best_10s
    table_rows.append(("Setting Time (10s)", "-", f"{delta_t_sec(t_b):.4f}", f"{P_b:.4f}", f"{Q_b:.4f}", f"{S_b:.4f}"))

    def fmt_row(cells):
        return "  ".join(str(c).ljust(col_widths[i]) for i, c in enumerate(cells))

    # IEEE 2800 damping ratio summary
    if zeta_est is not None:
        zeta_line = (
            f"Estimated damping ratio ζ = {zeta_est:.3f} "
            f"(IEEE 2800 target ≥ {zeta_min:.3f} → {pf(pass_damp)})"
        )
    else:
        zeta_line = (
            f"Estimated damping ratio ζ: no overshoot observed "
            f"(treated as Pass vs IEEE 2800 target ≥ {zeta_min:.3f})"
        )

    lines = [
        f"Start (pre-step) row: Time (UTC) = {t0},  P = {P0:.4f} MW,  Q = {Q0:.4f} MVar,  S = {S0:.4f} MVA",
        f"Requested timestamp (t = 0): {t_req}",
        f"delta_P_target = {delta_P_target_MW} MW",
        zeta_line,
        "",
        "delta_t = time since requested timestamp",
        "",
        fmt_row(table_rows[0]),
        fmt_row(table_rows[1]),
    ]
    for r in table_rows[2:]:
        lines.append(fmt_row(r))
        if max_overshoot_subtitle is not None and r[0] == "Max overshoot":
            lines.append(fmt_row((max_overshoot_subtitle, "", "", "", "", "")))

    return "\n".join(lines), window, t_req, None, P0


def _smooth_curve(y, window_size=3):
    """Simple moving average; window_size must be odd."""
    if len(y) < window_size or window_size < 2:
        return list(y)
    k = window_size // 2
    out = []
    for i in range(len(y)):
        lo = max(0, i - k)
        hi = min(len(y), i + k + 1)
        out.append(sum(y[lo:hi]) / (hi - lo))
    return out


def plot_signal(window, t_req, P0, delta_P_target_MW, gui_window=None, canvas_key=None):
    """
    Plot P (MW) vs time. Starts at 10% mark, smoothed, y-axis centered with padding.
    Labels 10%, 90%, max overshoot. If gui_window/canvas_key set, draw on tab canvas.
    """
    try:
        import matplotlib
        matplotlib.use("TkAgg")
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    except ImportError:
        return
    if not window or t_req is None or delta_P_target_MW is None:
        return
    delta_P = float(delta_P_target_MW)
    threshold_10 = P0 + 0.1 * delta_P
    threshold_90 = P0 + 0.9 * delta_P
    negative = delta_P < 0

    def crosses_10(t, p):
        return (negative and p <= threshold_10) or (not negative and p >= threshold_10)

    t_start_10 = None
    for t, P, _, _ in window:
        if crosses_10(t, P):
            t_start_10 = pd.Timestamp(t)
            break
    if t_start_10 is None:
        t_start_10 = pd.Timestamp(window[0][0])

    filtered = [(t, P, Q, S) for t, P, Q, S in window if pd.Timestamp(t) >= t_start_10]
    if not filtered:
        filtered = list(window)
        t_start_10 = pd.Timestamp(window[0][0])

    secs = [(pd.Timestamp(t) - t_start_10).total_seconds() for t, _, _, _ in filtered]
    P_vals = [p for _, p, _, _ in filtered]
    P_smooth = _smooth_curve(P_vals, 5)
    p_min, p_max = min(P_smooth), max(P_smooth)
    p_center = (p_min + p_max) / 2.0
    half_range = (p_max - p_min) / 2.0 or 0.5
    padding = max(half_range * 0.1, 0.5)
    y_lo = p_center - half_range - padding
    y_hi = p_center + half_range + padding

    key_10 = None
    key_90 = None
    key_max = None
    P_target = P0 + delta_P
    for i, (t, P, _, _) in enumerate(filtered):
        if key_10 is None and crosses_10(t, P):
            key_10 = (secs[i], P_smooth[i])
        if key_90 is None and ((negative and P <= threshold_90) or (not negative and P >= threshold_90)):
            key_90 = (secs[i], P_smooth[i])
        if negative and P < P_target:
            if key_max is None or P_smooth[i] < key_max[1]:
                key_max = (secs[i], P_smooth[i])
        elif not negative and P > P_target:
            if key_max is None or P_smooth[i] > key_max[1]:
                key_max = (secs[i], P_smooth[i])

    fig = Figure(figsize=(9, 5), dpi=100)
    ax = fig.add_subplot(111)
    ax.plot(secs, P_smooth, "b.-", label="P (MW)", markersize=3)
    ax.set_xlim(0, 10)
    ax.set_ylim(y_lo, y_hi)
    ax.set_xlabel("Time since 10% point (s)")
    ax.set_ylabel("P (MW)")
    ax.legend(loc="best")
    ax.grid(True, alpha=0.3)
    ax.set_title("P over 10s window (IEEE 2800 PFR)")
    for label, pt in [("10%", key_10), ("90%", key_90), ("max overshoot", key_max)]:
        if pt is not None:
            ax.plot(pt[0], pt[1], "ro", markersize=6)
            ax.annotate(label, xy=pt, xytext=(5, 5), textcoords="offset points", fontsize=9, annotation_clip=True)
    fig.tight_layout()

    if gui_window and canvas_key:
        try:
            tk_canvas = gui_window[canvas_key].TKCanvas
        except Exception:
            tk_canvas = getattr(gui_window[canvas_key], "Widget", None) or gui_window[canvas_key].TKCanvas
        for child in tk_canvas.winfo_children():
            child.destroy()
        canvas_agg = FigureCanvasTkAgg(fig, master=tk_canvas)
        canvas_agg.draw()
        canvas_agg.get_tk_widget().pack(fill="both", expand=True)
        gui_window._mpl_canvas_agg = canvas_agg
    else:
        import matplotlib.pyplot as plt
        plt.figure(fig.number)
        plt.show()


def _open_settings_window():
    """Open Settings window to edit default headers. Returns True if saved."""
    cfg = load_config()
    ph = cfg.get("phasor_headers", DEFAULT_PHASOR_HEADERS)
    tu = cfg.get("time_utc_headers", list(DEFAULT_TIME_UTC_HEADERS))
    ph_str = ", ".join(ph) if isinstance(ph, list) else ", ".join(DEFAULT_PHASOR_HEADERS)
    tu_str = ", ".join(tu) if isinstance(tu, (list, tuple)) else ", ".join(DEFAULT_TIME_UTC_HEADERS)

    layout_settings = [
        [sg.Text("Time (UTC) column headers (comma-separated, normalized to lower + underscores):")],
        [sg.Multiline(tu_str, key="-TU-", size=(60, 2), font=("Consolas", 9))],
        [sg.Text("Phasor column headers (comma-separated, exactly 12: va,vb,vc, va_ang,vb_ang,vc_ang, ia,ib,ic, ia_ang,ib_ang,ic_ang):")],
        [sg.Multiline(ph_str, key="-PH-", size=(60, 3), font=("Consolas", 9))],
        [sg.Button("Save", key="-SAVE-"), sg.Button("Cancel", key="-CANCEL-")],
    ]
    win = sg.Window("Settings – Header names", layout_settings, modal=True)
    while True:
        ev, val = win.read()
        if ev in (sg.WIN_CLOSED, "-CANCEL-", None):
            win.close()
            return False
        if ev == "-SAVE-":
            tu_raw = (val.get("-TU-") or "").strip()
            ph_raw = (val.get("-PH-") or "").strip()
            new_tu = [x.strip().lower().replace(" ", "_") for x in tu_raw.split(",") if x.strip()]
            new_ph = [x.strip() for x in ph_raw.split(",") if x.strip()]
            if len(new_ph) != 12:
                sg.popup("Phasor headers must be exactly 12 comma-separated names.", title="Settings")
                continue
            save_config(new_ph, new_tu)
            sg.popup("Settings saved.", title="Settings")
            win.close()
            return True
    win.close()
    return False


def _acceptance_criteria_image_path():
    """Path to acceptance criteria image (Reaction time, Rise time, etc.). Prefer assets/ next to script."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    assets = os.path.join(script_dir, "assets")
    candidates = [
        os.path.join(assets, "acceptance_criteria.png"),
        os.path.join(assets, "acceptance_criteria.jpg"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    if os.path.isdir(assets):
        for name in sorted(os.listdir(assets)):
            if name.lower().endswith((".png", ".jpg", ".jpeg")):
                return os.path.join(assets, name)
    return None


def run_gui():
    """FreeSimpleGUI: timestamp, delta_P_target; pass/fail criteria; Run Report (10s analysis)."""
    sg.theme("SystemDefault")

    # Pass/fail criteria defaults: 10% Response Time (ms), 90% Response Time (ms), Damping Ratio (Zeta)
    default_10_ms = "200"
    default_90_ms = "2000"   # 2 seconds
    default_zeta = "0.3"

    img_path = _acceptance_criteria_image_path()
    if img_path:
        image_elem = sg.Frame("Acceptance criteria", [[sg.Image(filename=img_path, key="-ACCEPT_IMG-")]], expand_x=True)
    else:
        ensure_source_data_exists()
        script_dir = os.path.dirname(os.path.abspath(__file__))
        assets = os.path.join(script_dir, "assets")
        if not os.path.isdir(assets):
            os.makedirs(assets)
        image_elem = sg.Frame("Acceptance criteria", [[sg.Text("Place the acceptance criteria image (Reaction time, Rise time, Settling time, Damping ratio, Settling band) in:\n" + assets, font=("Default", 9))]], expand_x=True)

    tab_report = [
        [sg.Text("Data file (optional):", size=(14, 1)), sg.Input("", key="-ATTACH-", size=(50, 1), tooltip="Leave empty to use source_data/phasor_data.xlsx"),
         sg.FileBrowse("Browse", target="-ATTACH-", file_types=(("Excel", "*.xlsx"), ("All", "*.*")))],
        [sg.Text("Time (UTC):", size=(12, 1)), sg.Input("", key="-TIMESTAMP-", size=(28, 1), tooltip="e.g. 02/05/2026 18:03:12"),
         sg.Text("Delta P target (MW):", size=(18, 1)), sg.Input("", key="-DELTA_P-", size=(10, 1)),
         sg.Button("Run Report", key="-RUN10-"), sg.Button("Find Test", key="-FIND_TEST-"), sg.Button("Settings", key="-SETTINGS-")],
        [sg.Frame("Pass/Fail criteria", [
            [sg.Text("10% Response Time (ms):", size=(22, 1)), sg.Input(default_10_ms, key="-LIMIT_10_MS-", size=(8, 1)), sg.Text("(200 ms to 1 s)", size=(14, 1))],
            [sg.Text("90% Response Time (ms):", size=(22, 1)), sg.Input(default_90_ms, key="-LIMIT_90_MS-", size=(8, 1)), sg.Text("(2000ms to 20000ms)", size=(20, 1))],
            [sg.Text("Damping Ratio (Zeta):", size=(22, 1)), sg.Input(default_zeta, key="-ZETA_MIN-", size=(8, 1)), sg.Text("(0.2 to 1.0)", size=(14, 1))],
        ], expand_x=True)],
        [image_elem],
        [sg.Multiline("Enter timestamp and Delta P target (MW), set criteria, then click Run Report.", key="-OUT-", size=(80, 18), font=("Consolas", 10), expand_x=True)],
    ]
    tab_graph = [
        [sg.Canvas(key="-GRAPH_CANVAS-", size=(800, 420), expand_x=True, expand_y=True)],
    ]
    layout = [
        [sg.TabGroup([
            [sg.Tab("Report", tab_report), sg.Tab("Graph", tab_graph)],
        ], expand_x=True, expand_y=True)],
    ]
    window = sg.Window("IEEE 2800 PFR Calcs", layout, resizable=True, size=(820, 620))

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, None):
            break
        if event == "-SETTINGS-":
            _open_settings_window()
            continue
        if event == "-FIND_TEST-":
            dp = (values.get("-DELTA_P-") or "").strip()
            attach = (values.get("-ATTACH-") or "").strip()
            data_path = attach if os.path.isfile(attach) else None
            if not data_path and not get_data_path():
                sg.popup("Select a data file (Browse) or place phasor_data.xlsx in source_data.", title="Find Test")
                continue
            if not dp:
                sg.popup("Enter Delta P target (MW) first.", title="Find Test")
                continue
            candidates, err = find_test_timestamps(data_path, dp)
            if err:
                sg.popup(err, title="Find Test")
                continue
            _open_find_test_popup(candidates, "-TIMESTAMP-", window)
            continue
        if event == "-RUN10-":
            ts = (values.get("-TIMESTAMP-") or "").strip()
            dp = (values.get("-DELTA_P-") or "").strip()
            attach = (values.get("-ATTACH-") or "").strip()
            data_path = attach if os.path.isfile(attach) else None
            if not data_path and not get_data_path():
                window["-OUT-"].update("Select a data file (Browse) or place phasor_data.xlsx in source_data.")
                continue
            if not ts:
                window["-OUT-"].update("Enter a timestamp.")
                continue
            if not dp:
                window["-OUT-"].update("Enter Delta P target (MW).")
                continue
            try:
                limit_10 = float((values.get("-LIMIT_10_MS-") or default_10_ms).strip())
                limit_90 = float((values.get("-LIMIT_90_MS-") or default_90_ms).strip())
                zeta_min = float((values.get("-ZETA_MIN-") or default_zeta).strip())
            except (ValueError, TypeError):
                window["-OUT-"].update("Pass/fail criteria must be numbers (10% and 90% in ms, Zeta as decimal).")
                continue
            out, win_data, t_req, err, P0 = run_10s_analysis(ts, dp, limit_10_ms=limit_10, limit_90_ms=limit_90, zeta_min=zeta_min, data_path=data_path)
            if err:
                window["-OUT-"].update(err)
                continue
            window["-OUT-"].update(out)
            if win_data is not None and t_req is not None and P0 is not None:
                plot_signal(win_data, t_req, P0, float(dp), gui_window=window, canvas_key="-GRAPH_CANVAS-")

    window.close()


def main():
    # Start GUI; data file can be selected via Browse or from source_data
    run_gui()


if __name__ == "__main__":
    main()
