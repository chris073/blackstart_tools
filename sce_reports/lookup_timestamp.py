"""One-off: lookup 2/5/2026 6:03:12 PM in phasor_data and print phasor values + P,Q,S calc. Also show row 2155."""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from sce_reports import get_data_path, read_phasor_row_at_timestamp, calc_pqs_from_phasor_row, PHASOR_HEADERS

TS = "2026-02-05 18:03:12"  # 2/5/2026 6:03:12 PM
CHECK_ROW = 1255  # Excel line 1255 (row 1254 in 0-based data)

def main():
    path = get_data_path()
    print(f"Data file: {path}")
    if not path or not os.path.exists(path):
        print("File not found.")
        return
    result, err = read_phasor_row_at_timestamp(TS)
    if err:
        print(f"Error: {err}")
        return
    headers, row_vals, time_actual, row_num = result
    print(f"\n--- Lookup by timestamp '{TS}' -> Excel row {row_num} ---")
    print(f"Time (UTC): {time_actual}")
    print("Phasor data:")
    for h, v in zip(headers, row_vals):
        print(f"  {h}: {v}")
    P_MW, Q_MVar, S_MVA = calc_pqs_from_phasor_row(row_vals)
    print(f"P = {P_MW:.4f} MW,  Q = {Q_MVar:.4f} MVar,  S = {S_MVA:.4f} MVA")

    # Also read row 2155 directly to confirm
    if path.endswith(".xlsx"):
        import openpyxl
        from sce_reports import SHEET_NAME, _normalize_header, TIME_UTC_HEADERS
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        for name in wb.sheetnames:
            if (name or "").strip().lower() == SHEET_NAME.strip().lower():
                ws = wb[name]
                break
        else:
            ws = None
        if ws and CHECK_ROW <= ws.max_row:
            header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            normalized = [_normalize_header(h) for h in header_row]
            time_col = None
            for n in TIME_UTC_HEADERS:
                try:
                    idx = normalized.index(n)
                    time_col = idx + 1
                    break
                except ValueError:
                    continue
            col_indices = []
            for name in PHASOR_HEADERS:
                try:
                    idx = normalized.index(name)
                    col_indices.append(idx + 1)
                except ValueError:
                    break
            if time_col and len(col_indices) == 12:
                t_val = ws.cell(row=CHECK_ROW, column=time_col).value
                row_vals_2155 = [ws.cell(row=CHECK_ROW, column=col).value for col in col_indices]
                print(f"\n--- Row {CHECK_ROW} (direct read) ---")
                print(f"Time (UTC): {t_val}")
                print("Phasor data:")
                for h, v in zip(PHASOR_HEADERS, row_vals_2155):
                    print(f"  {h}: {v}")
                try:
                    P2, Q2, S2 = calc_pqs_from_phasor_row(row_vals_2155)
                    print(f"P = {P2:.4f} MW,  Q = {Q2:.4f} MVar,  S = {S2:.4f} MVA")
                except Exception as e:
                    print(f"Calc error: {e}")
        if ws:
            wb.close()

if __name__ == "__main__":
    main()
