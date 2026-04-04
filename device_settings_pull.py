import FreeSimpleGUI as sg
import requests
import pandas as pd
import json
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Load .env file if it exists
load_dotenv()

def flatten_json(d, parent_key='', sep='.'):
    items = []
    if not isinstance(d, dict):
        return {"data": str(d)}
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_json(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            for i, item in enumerate(v):
                if isinstance(item, (dict, list)):
                    items.extend(flatten_json({str(i): item}, new_key, sep=sep).items())
                else:
                    items.append((f"{new_key}{sep}{i}", item))
        else:
            items.append((new_key, v))
    return dict(items)

def create_excel(data, filename, customer_name, device_id):
    # Flatten Data
    if isinstance(data, list):
        df = pd.DataFrame([flatten_json(item) for item in data])
    else:
        flat_dict = flatten_json(data)
        df = pd.DataFrame(list(flat_dict.items()), columns=['Setting Name', 'Setting Value'])

    # Sheet name is the device ID (Excel limit 31 chars)
    sheet_name = str(device_id)[:31]

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, startcol=1, sheet_name=sheet_name)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Styles
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, size=12)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Header: Customer & Date
        worksheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
        top_cell = worksheet.cell(row=2, column=2)
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        top_cell.value = f"Customer: {customer_name} | Generated: {date_str}"
        top_cell.fill = title_fill
        top_cell.font = title_font
        top_cell.alignment = Alignment(horizontal="center")

        # Table Column Headers
        for col_num in range(2, 2 + len(df.columns)): 
            cell = worksheet.cell(row=4, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            worksheet.column_dimensions[cell.column_letter].width = 45

            # Table Grid Lines
            for row_num in range(5, 5 + len(df)):
                worksheet.cell(row=row_num, column=col_num).border = thin_border

def make_request(url, api_key, customer, device):
    headers = {"X-API-KEY": api_key, "Content-Type": "application/json"}
    payload = {"customerName": customer, "deviceId": device}
    response = requests.post(url, json=payload, headers=headers, timeout=15)
    response.raise_for_status()
    return response.json()

# --- GUI ---
endpoint_map = {
    "Device Config": "deviceconfig",
    "Fleet Config": "fleetconfiginternal",
    "Devices": "devicesinternal"
}

# Initial value for API key from .env
env_key = os.getenv("X_API_KEY", "")
default_url = "https://f6xfmd43n3.execute-api.us-west-2.amazonaws.com/test"

layout = [
    [sg.Text("API Config Manager", font=("Helvetica", 18, "bold"))],
    [sg.Text("Base URL:"), sg.Input(default_url, key="-URL-", expand_x=True)],
    [sg.Text("API Key:"), sg.Input(env_key, key="-KEY-", password_char="*", expand_x=True)],
    [sg.Frame("Endpoint", [[sg.Radio(k, "EP", key=k, default=(k=="Device Config")) for k in endpoint_map.keys()]])],
    [sg.Text("Customer Name:"), sg.Input("indie", key="-CUST-", expand_x=True)],
    [sg.Text("Device ID:"), sg.Input("XXX_XXX_U_01", key="-DEVICE-", expand_x=True)],
    [sg.Button("Test Connection", button_color="#2980b9"), 
     sg.Button("Export to Excel", key="-EXPORT-", disabled=True, button_color="green"), 
     sg.Button("Exit")],
    [sg.Multiline(size=(85, 12), key="-OUTPUT-", font=("Courier New", 10), background_color="#1e1e1e", text_color="#dcdcdc")],
    [sg.StatusBar("Ready", key="-STATUS-")]
]

window = sg.Window("API Config Tool", layout, finalize=True)
latest_data = None 

while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED, "Exit"):
        break
    
    selected_label = next((k for k in endpoint_map if values[k]), "Device Config")
    full_url = f"{values['-URL-'].rstrip('/')}/{endpoint_map[selected_label]}"
    
    if event == "Test Connection":
        window["-OUTPUT-"].update("Testing...")
        try:
            latest_data = make_request(full_url, values["-KEY-"], values["-CUST-"], values["-DEVICE-"])
            window["-OUTPUT-"].update(f">>> SUCCESS\n{json.dumps(latest_data, indent=4)}")
            window["-EXPORT-"].update(disabled=False)
            window["-STATUS-"].update("Success. Ready to export.")
        except Exception as e:
            window["-OUTPUT-"].update(f">>> ERROR: {e}")
            window["-EXPORT-"].update(disabled=True)

    if event == "-EXPORT-":
        file_path = sg.popup_get_file('Save As', save_as=True, default_extension=".xlsx", 
                                    default_path=f"{values['-CUST-']}_{values['-DEVICE-']}.xlsx")
        if file_path:
            create_excel(latest_data, file_path, values["-CUST-"], values["-DEVICE-"])
            sg.popup("Export Complete", f"Sheet name set to: {values['-DEVICE-']}")

window.close()